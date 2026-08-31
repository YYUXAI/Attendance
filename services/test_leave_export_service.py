from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

from services.leave_export_service import (
    LeaveExportRow,
    encode_leave_export_xlsx,
    summarize_leave_rows_by_person,
)


def _row(
    *,
    name: str,
    eid: str,
    leave_hm: str,
    back_hm: str = "",
    duration_minutes: int | None = None,
) -> LeaveExportRow:
    duration_text = f"{duration_minutes}分" if duration_minutes is not None else ""
    return LeaveExportRow(
        english_name=name,
        employee_id=eid,
        leave_hm=leave_hm,
        back_hm=back_hm,
        duration_text=duration_text,
        reason="",
        duration_minutes=duration_minutes,
    )


def test_summarize_leave_rows_by_person_for_today_range() -> None:
    rows = [
        _row(name="singjang", eid="56773", leave_hm="14:00", back_hm="14:20", duration_minutes=20),
        _row(name="singjang", eid="56773", leave_hm="15:00", back_hm="15:40", duration_minutes=40),
        _row(name="singjang", eid="56773", leave_hm="16:00", back_hm="16:30", duration_minutes=30),
        _row(name="A", eid="1", leave_hm="09:00"),
    ]
    summary = summarize_leave_rows_by_person(rows=rows, overtime_minutes=30)
    by_id = {item.employee_id: item for item in summary}
    assert by_id["56773"].english_name == "singjang"
    assert by_id["56773"].leave_count == 3
    assert by_id["56773"].return_count == 3
    assert by_id["56773"].overtime_count == 2
    assert by_id["1"].leave_count == 1
    assert by_id["1"].return_count == 0
    assert by_id["1"].overtime_count == 0


def test_leave_export_xlsx_includes_summary_sheet() -> None:
    rows = [
        _row(name="singjang", eid="56773", leave_hm="10:00", back_hm="10:40", duration_minutes=40),
        _row(name="singjang", eid="56773", leave_hm="11:00", back_hm="11:10", duration_minutes=10),
        _row(name="singjang", eid="56773", leave_hm="12:00", back_hm="12:20", duration_minutes=20),
    ]
    payload = encode_leave_export_xlsx(rows=rows, overtime_minutes=30)
    wb = load_workbook(BytesIO(payload))
    assert wb.sheetnames == ["离岗返岗", "汇总"]
    ws = wb["汇总"]
    assert [cell.value for cell in ws[1]] == [
        "姓名",
        "工号",
        "离岗次数",
        "返岗次数",
        "超过30分钟次数",
    ]
    assert [cell.value for cell in ws[2]] == ["singjang", "56773", 3, 3, 1]
