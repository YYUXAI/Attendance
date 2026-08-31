"""T-上班报备群等：离岗返岗记录导出。"""
from __future__ import annotations

import io
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta, timezone
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from repositories import temporary_leave_records_repo

_SH = ZoneInfo("Asia/Shanghai")
_HEADERS = ("姓名", "工号", "离岗", "返岗", "时长", "原因")
_SUMMARY_SHEET_TITLE = "汇总"


@dataclass(frozen=True)
class LeaveExportRow:
    english_name: str
    employee_id: str
    leave_hm: str
    back_hm: str
    duration_text: str
    reason: str
    duration_minutes: int | None = None


@dataclass(frozen=True)
class LeavePersonSummary:
    english_name: str
    employee_id: str
    leave_count: int
    return_count: int
    overtime_count: int


def shanghai_day_bounds_utc(*, start: date, end: date) -> tuple[datetime, datetime]:
    start_local = datetime.combine(start, time.min, tzinfo=_SH)
    end_exclusive_local = datetime.combine(end + timedelta(days=1), time.min, tzinfo=_SH)
    return start_local.astimezone(timezone.utc), end_exclusive_local.astimezone(timezone.utc)


def collect_leave_rows_for_chat(
    *,
    chat_id: int,
    start: date,
    end: date,
) -> list[LeaveExportRow]:
    start_utc, end_exclusive_utc = shanghai_day_bounds_utc(start=start, end=end)
    records = temporary_leave_records_repo.list_by_chat_and_range(
        chat_id=int(chat_id),
        start_utc=start_utc,
        end_utc=end_exclusive_utc,
    )
    filtered: list[tuple[datetime, LeaveExportRow]] = []
    for rec in records:
        leave_at = _as_shanghai(rec.leave_at)
        if leave_at is None:
            continue
        leave_day = leave_at.date()
        if leave_day < start or leave_day > end:
            continue
        back_at = _as_shanghai(rec.back_at)
        duration = rec.duration_minutes
        if duration is None and back_at is not None:
            duration = max(0, int((back_at - leave_at).total_seconds() // 60))
        duration_int = int(duration) if duration is not None else None
        filtered.append(
            (
                leave_at,
                LeaveExportRow(
                    english_name=(rec.english_name or "").strip() or "未命名",
                    employee_id=str(rec.employee_id).strip(),
                    leave_hm=leave_at.strftime("%H:%M"),
                    back_hm=back_at.strftime("%H:%M") if back_at is not None else "",
                    duration_text=f"{duration_int}分" if duration_int is not None else "",
                    reason=(rec.reason or "").strip(),
                    duration_minutes=duration_int,
                ),
            )
        )
    filtered.sort(key=lambda item: item[0], reverse=True)
    return [row for _, row in filtered]


def summarize_leave_rows_by_person(
    *,
    rows: list[LeaveExportRow],
    overtime_minutes: int,
) -> list[LeavePersonSummary]:
    """按工号汇总导出范围内的离岗 / 返岗 / 超时次数。"""
    threshold = int(overtime_minutes)
    grouped: dict[str, LeavePersonSummary] = {}
    order: list[str] = []
    for row in rows:
        eid = str(row.employee_id).strip()
        if not eid:
            continue
        if eid not in grouped:
            order.append(eid)
            grouped[eid] = LeavePersonSummary(
                english_name=(row.english_name or "").strip() or "未命名",
                employee_id=eid,
                leave_count=0,
                return_count=0,
                overtime_count=0,
            )
        current = grouped[eid]
        overtime = (
            row.duration_minutes is not None and int(row.duration_minutes) >= threshold
        )
        grouped[eid] = LeavePersonSummary(
            english_name=current.english_name,
            employee_id=eid,
            leave_count=current.leave_count + 1,
            return_count=current.return_count + (1 if (row.back_hm or "").strip() else 0),
            overtime_count=current.overtime_count + (1 if overtime else 0),
        )
    return [grouped[eid] for eid in sorted(order)]


def encode_leave_export_xlsx(
    *,
    rows: list[LeaveExportRow],
    overtime_minutes: int = 30,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "离岗返岗"
    header_fill = PatternFill("solid", fgColor="1F1F1F")
    header_font = Font(color="FFFFFF", bold=True)
    body_font = Font(color="FFFFFF")
    body_fill = PatternFill("solid", fgColor="2B2B2B")
    _write_header(ws, _HEADERS, header_fill=header_fill, header_font=header_font)
    for row_idx, row in enumerate(rows, start=2):
        values = (
            row.english_name,
            row.employee_id,
            row.leave_hm,
            row.back_hm,
            row.duration_text,
            row.reason,
        )
        _write_body_row(
            ws,
            row_idx,
            values,
            body_fill=body_fill,
            body_font=body_font,
        )
    widths = (16, 12, 10, 10, 10, 24)
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width

    summary_headers = (
        "姓名",
        "工号",
        "离岗次数",
        "返岗次数",
        f"超过{int(overtime_minutes)}分钟次数",
    )
    summary = summarize_leave_rows_by_person(
        rows=rows,
        overtime_minutes=overtime_minutes,
    )
    ws_sum = wb.create_sheet(_SUMMARY_SHEET_TITLE)
    _write_header(ws_sum, summary_headers, header_fill=header_fill, header_font=header_font)
    for row_idx, person in enumerate(summary, start=2):
        _write_body_row(
            ws_sum,
            row_idx,
            (
                person.english_name,
                person.employee_id,
                person.leave_count,
                person.return_count,
                person.overtime_count,
            ),
            body_fill=body_fill,
            body_font=body_font,
        )
    summary_widths = (16, 12, 12, 12, 18)
    for idx, width in enumerate(summary_widths, start=1):
        ws_sum.column_dimensions[chr(64 + idx)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def leave_export_filename(*, start: date, end: date) -> str:
    if start == end:
        return f"离岗返岗_{start.isoformat()}.xlsx"
    return f"离岗返岗_{start.isoformat()}_{end.isoformat()}.xlsx"


def _write_header(ws, headers: tuple[str, ...], *, header_fill, header_font) -> None:
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _write_body_row(ws, row_idx: int, values: tuple[object, ...], *, body_fill, body_font) -> None:
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col, value=value)
        cell.fill = body_fill
        cell.font = body_font
        cell.alignment = Alignment(horizontal="left", vertical="center")


def _as_shanghai(value: object) -> datetime | None:
    if not isinstance(value, datetime):
        return None
    if value.tzinfo is None:
        value = value.replace(tzinfo=timezone.utc)
    return value.astimezone(_SH)
