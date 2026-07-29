from __future__ import annotations

import io
import logging
import os
from dataclasses import dataclass, field, replace
from datetime import date, timedelta
from typing import Dict, Iterable, List, Literal, Optional, Set
from zoneinfo import ZoneInfo

from aiogram import Bot

from infra.attendance_override_config import load_attendance_override_config
from infra.bbq_google_sheets_config import bbq_summary_excluded_employee_ids
from infra.checkin_remote_diff_config import yymg_summary_excluded_employee_ids
from infra.google_sheets_alt_config import load_google_sheets_alt_config
from infra.db import get_cursor
from repositories import registrations_repo
from services.shift_view_service import (
    SHIFT_VIEW_ALT,
    SHIFT_VIEW_MAIN,
    roster_employee_ids,
    shift_view_for_tg_id,
)
from services.group_attendance_summary_service import (
    AttendanceSummaryRow,
    build_rows_for_group,
    build_rows_for_monthly_roster_remainder,
    build_rows_for_roster_at_chat,
    dedupe_export_rows_by_employee,
    list_attendance_group_ids,
    resolve_group_display_name,
    fallback_group_display_name_from_db,
)

log = logging.getLogger(__name__)

ExportRangeKind = Literal[
    "today",
    "yesterday",
    "week",
    "last_week",
    "month",
    "last_month",
]

_ABNORMAL_STATUSES = frozenset({"缺勤", "缺卡", "迟到", "早退", "迟到+早退"})

# 供 Google 表等复用
ABNORMAL_EXPORT_STATUSES = _ABNORMAL_STATUSES

# 状态分布图配色（不含应出勤/实出勤人数）
_STATUS_CHART_COLORS: tuple[tuple[str, str], ...] = (
    ("正常", "A5A5A5"),
    ("月休", "5B9BD5"),
    ("缺勤", "FFC000"),
    ("迟到", "ED7D31"),
    ("早退", "70AD47"),
    ("缺卡", "C00000"),
    ("请假", "7030A0"),
)


def _export_scope_chat_ids() -> frozenset[int]:
    """
    导出群范围：
    - 未配置 CHECKIN_EXPORT_CHAT_IDS：导出全量考勤群（历史行为）
    - 已配置：仅导出指定群，其他群不计入统计
    """
    raw = (os.getenv("CHECKIN_EXPORT_CHAT_IDS") or "").strip()
    if not raw:
        return frozenset()
    out: set[int] = set()
    for part in raw.replace("，", ",").split(","):
        p = part.strip()
        if not p:
            continue
        try:
            out.add(int(p))
        except ValueError:
            continue
    return frozenset(out)


def _filter_bbq_export_excluded(rows: List[AttendanceSummaryRow]) -> List[AttendanceSummaryRow]:
    excluded = bbq_summary_excluded_employee_ids()
    if not excluded:
        return rows
    return [r for r in rows if str(r.employee_id).strip() not in excluded]


def _registered_english_name(*, employee_id: str) -> str | None:
    reg = registrations_repo.get_by_employee_id(str(employee_id).strip())
    if not reg:
        return None
    name = (reg.english_name or "").strip()
    return name or None


def _with_registered_name(row: AttendanceSummaryRow) -> AttendanceSummaryRow:
    name = _registered_english_name(employee_id=str(row.employee_id))
    if name:
        return replace(row, english_name=name)
    return row


def _mirror_attendance_content(
    *,
    src_row: AttendanceSummaryRow,
    dst_employee_id: str,
    target_date: date,
) -> AttendanceSummaryRow:
    """复制打卡/状态等内容，姓名与工号保留目标员工本人。"""
    dst_name = _registered_english_name(employee_id=dst_employee_id) or src_row.english_name
    return replace(
        src_row,
        employee_id=str(dst_employee_id),
        english_name=dst_name,
        work_date=target_date,
    )


async def _apply_export_employee_overrides(
    *,
    rows: List[AttendanceSummaryRow],
    target_date: date,
    bot: Bot | None,
) -> List[AttendanceSummaryRow]:
    cfg = load_attendance_override_config()
    if not cfg.export_employee_chat and not cfg.export_mirror_from:
        return rows

    by_eid: Dict[str, AttendanceSummaryRow] = {}
    for r in rows:
        eid = str(r.employee_id).strip()
        if eid:
            by_eid[eid] = r

    override_eids = set(cfg.export_employee_chat) | set(cfg.export_mirror_from)
    for eid in override_eids:
        by_eid.pop(eid, None)

    async def _row_from_chat(*, eid: str, chat_id: int) -> AttendanceSummaryRow | None:
        if bot is not None:
            gname = await resolve_group_display_name(bot=bot, chat_id=int(chat_id))
        else:
            gname = fallback_group_display_name_from_db(chat_id=int(chat_id))
        day_rows = build_rows_for_group(
            chat_id=int(chat_id),
            target_date=target_date,
            group_name=gname,
        )
        picked = next((r for r in day_rows if str(r.employee_id).strip() == eid), None)
        return replace(picked, work_date=target_date) if picked else None

    for eid, chat_id in cfg.export_employee_chat.items():
        if eid in yymg_summary_excluded_employee_ids():
            continue
        picked = await _row_from_chat(eid=eid, chat_id=int(chat_id))
        if picked:
            by_eid[eid] = _with_registered_name(picked)

    for dst, src in cfg.export_mirror_from.items():
        if dst in yymg_summary_excluded_employee_ids():
            continue
        src_row = by_eid.get(src)
        if not src_row:
            chat_id = cfg.export_employee_chat.get(src)
            if chat_id is not None:
                src_row = await _row_from_chat(eid=src, chat_id=int(chat_id))
                if src_row:
                    src_row = _with_registered_name(src_row)
        if src_row:
            by_eid[dst] = _mirror_attendance_content(
                src_row=src_row,
                dst_employee_id=dst,
                target_date=target_date,
            )

    return list(by_eid.values())


@dataclass(frozen=True)
class AttendanceExportOverview:
    expected_count: int
    actual_count: int
    monthly_rest: int
    absent: int
    late: int
    early: int
    missed_punch: int
    leave: int


@dataclass
class EmployeeExportPivot:
    group_name: str
    english_name: str
    employee_id: str
    daily_status: Dict[date, str] = field(default_factory=dict)


def resolve_export_date_range(
    *,
    kind: ExportRangeKind,
    today: date,
) -> tuple[date, date, str]:
    if kind == "today":
        return today, today, "今日"
    if kind == "yesterday":
        d = today - timedelta(days=1)
        return d, d, "昨天"
    if kind == "week":
        start = today - timedelta(days=today.weekday())
        end = min(today, start + timedelta(days=6))
        return start, end, "本周"
    if kind == "last_week":
        this_week_start = today - timedelta(days=today.weekday())
        end = this_week_start - timedelta(days=1)
        start = end - timedelta(days=6)
        return start, end, "上周"
    if kind == "month":
        start = date(today.year, today.month, 1)
        return start, today, "本月"
    # last_month: 上一个自然月整月
    first_this_month = date(today.year, today.month, 1)
    end = first_this_month - timedelta(days=1)
    start = date(end.year, end.month, 1)
    return start, end, "上月"


def export_filename(*, start: date, end: date) -> str:
    if start == end:
        return f"{start.isoformat()}.xlsx"
    return f"{start.isoformat()}_{end.isoformat()}.xlsx"


def _date_range_inclusive(*, start: date, end: date) -> List[date]:
    out: List[date] = []
    d = start
    while d <= end:
        out.append(d)
        d += timedelta(days=1)
    return out


def _date_column_label(d: date) -> str:
    return f"{d.month}月{d.day}日"


def _fetch_leave_dates_in_range(*, start: date, end: date) -> Dict[str, Set[date]]:
    with get_cursor() as cur:
        cur.execute(
            """
            SELECT employee_id, leave_date
            FROM public.effective_leave_days
            WHERE leave_date >= %s AND leave_date <= %s
            """,
            (start, end),
        )
        rows = cur.fetchall() or []
    out: Dict[str, Set[date]] = {}
    for eid, ld in rows:
        key = str(eid).strip()
        if not key or ld is None:
            continue
        out.setdefault(key, set()).add(ld)
    return out


def normalize_export_status(
    row: AttendanceSummaryRow,
    *,
    on_leave: bool,
) -> str:
    """导出展示口径：缺勤=上下班皆无；缺卡=缺一侧打卡。"""
    if on_leave:
        return "请假"
    st = (row.status or "").strip()
    if st == "月休":
        return "月休"
    has_in = bool((row.first_clock_local or "").strip())
    has_out = bool((row.last_clock_local or "").strip())
    if not has_in and not has_out:
        return "缺勤"
    if st == "正常":
        return "正常"
    if st == "迟到+早退":
        return "迟到+早退"
    if st in ("迟到", "早退"):
        return st
    if has_in != has_out:
        return "缺卡"
    return st or "缺勤"


async def collect_rows_for_date(
    *,
    target_date: date,
    bot: Bot | None,
    export_tg_id: int | None = None,
) -> List[AttendanceSummaryRow]:
    ym = target_date.strftime("%Y-%m")
    view = shift_view_for_tg_id(tg_id=int(export_tg_id)) if export_tg_id else SHIFT_VIEW_MAIN
    roster_ids = roster_employee_ids(year_month=ym, view=view) if export_tg_id else frozenset()

    # MGZ 班表管理员：按 Google 班表 roster，打卡仅统计 YYMG 考勤群
    if export_tg_id and view == SHIFT_VIEW_ALT and roster_ids:
        alt_cfg = load_google_sheets_alt_config()
        alt_chat = alt_cfg.attendance_chat_id if alt_cfg else None
        if alt_chat is not None:
            if bot is not None:
                gname = await resolve_group_display_name(bot=bot, chat_id=int(alt_chat))
            else:
                gname = fallback_group_display_name_from_db(chat_id=int(alt_chat))
            deduped = build_rows_for_roster_at_chat(
                target_date=target_date,
                chat_id=int(alt_chat),
                only_employee_ids=roster_ids,
                group_name=gname,
            )
        else:
            deduped = build_rows_for_monthly_roster_remainder(
                target_date=target_date,
                exclude_employee_ids=set(),
                only_employee_ids=roster_ids,
            )
        deduped = dedupe_export_rows_by_employee(rows=deduped)
        deduped = await _apply_export_employee_overrides(
            rows=deduped,
            target_date=target_date,
            bot=bot,
        )
        excluded = yymg_summary_excluded_employee_ids()
        deduped = [
            r
            for r in deduped
            if str(r.employee_id).strip() in roster_ids
            and str(r.employee_id).strip() not in excluded
        ]
        return [replace(r, work_date=target_date) for r in deduped]

    group_ids = list_attendance_group_ids()
    scoped_ids = _export_scope_chat_ids()
    if scoped_ids:
        group_ids = [gid for gid in group_ids if int(gid) in scoped_ids]
        log.info("attendance_export: scoped groups=%s", sorted(group_ids))
    all_rows: List[AttendanceSummaryRow] = []
    for gid in group_ids:
        if bot is not None:
            gname = await resolve_group_display_name(bot=bot, chat_id=int(gid))
        else:
            gname = fallback_group_display_name_from_db(chat_id=int(gid))
        all_rows.extend(
            build_rows_for_group(
                chat_id=int(gid),
                target_date=target_date,
                group_name=gname,
            )
        )
    seen = {str(r.employee_id).strip() for r in all_rows}
    if export_tg_id and roster_ids:
        all_rows.extend(
            build_rows_for_monthly_roster_remainder(
                target_date=target_date,
                exclude_employee_ids=seen,
                only_employee_ids=roster_ids,
            )
        )
    elif not scoped_ids:
        all_rows.extend(
            build_rows_for_monthly_roster_remainder(
                target_date=target_date,
                exclude_employee_ids=seen,
            )
        )
    deduped = dedupe_export_rows_by_employee(rows=all_rows)
    deduped = await _apply_export_employee_overrides(
        rows=deduped,
        target_date=target_date,
        bot=bot,
    )
    if export_tg_id and roster_ids:
        deduped = [
            r for r in deduped if str(r.employee_id).strip() in roster_ids
        ]
        present = {str(r.employee_id).strip() for r in deduped}
        missing = roster_ids - present
        if missing:
            deduped.extend(
                build_rows_for_monthly_roster_remainder(
                    target_date=target_date,
                    exclude_employee_ids=present,
                    only_employee_ids=missing,
                )
            )
            deduped = dedupe_export_rows_by_employee(rows=deduped)
            deduped = await _apply_export_employee_overrides(
                rows=deduped,
                target_date=target_date,
                bot=bot,
            )
            deduped = [
                r for r in deduped if str(r.employee_id).strip() in roster_ids
            ]
    if view == SHIFT_VIEW_MAIN:
        deduped = _filter_bbq_export_excluded(deduped)
    return [replace(r, work_date=target_date) for r in deduped]


async def collect_rows_for_range(
    *,
    start: date,
    end: date,
    bot: Bot | None,
    export_tg_id: int | None = None,
) -> List[AttendanceSummaryRow]:
    out: List[AttendanceSummaryRow] = []
    for d in _date_range_inclusive(start=start, end=end):
        out.extend(
            await collect_rows_for_date(
                target_date=d,
                bot=bot,
                export_tg_id=export_tg_id,
            )
        )
    return out


async def collect_rows_for_single_group(
    *,
    chat_id: int,
    start: date,
    end: date,
    bot: Bot | None,
) -> List[AttendanceSummaryRow]:
    """仅收集指定群的考勤行（用于工号打卡群 Google 表同步等）。"""
    out: List[AttendanceSummaryRow] = []
    for d in _date_range_inclusive(start=start, end=end):
        if bot is not None:
            gname = await resolve_group_display_name(bot=bot, chat_id=int(chat_id))
        else:
            gname = fallback_group_display_name_from_db(chat_id=int(chat_id))
        day_rows = build_rows_for_group(
            chat_id=int(chat_id),
            target_date=d,
            group_name=gname,
        )
        out.extend(replace(r, work_date=d) for r in day_rows)
    return out


async def collect_rows_for_roster_at_chat(
    *,
    chat_id: int,
    start: date,
    end: date,
    roster_ids: frozenset[str],
    bot: Bot | None,
) -> List[AttendanceSummaryRow]:
    """按班表 roster 全员统计；打卡/离岗记录仅取自指定群。"""
    if not roster_ids:
        return []
    out: List[AttendanceSummaryRow] = []
    for d in _date_range_inclusive(start=start, end=end):
        if bot is not None:
            gname = await resolve_group_display_name(bot=bot, chat_id=int(chat_id))
        else:
            gname = fallback_group_display_name_from_db(chat_id=int(chat_id))
        day_rows = build_rows_for_roster_at_chat(
            target_date=d,
            chat_id=int(chat_id),
            only_employee_ids=roster_ids,
            group_name=gname,
        )
        out.extend(replace(r, work_date=d) for r in day_rows)
    return out


def build_pivot_and_overview(
    *,
    rows: Iterable[AttendanceSummaryRow],
    start: date,
    end: date,
) -> tuple[List[EmployeeExportPivot], AttendanceExportOverview, List[date]]:
    dates = _date_range_inclusive(start=start, end=end)
    leave_map = _fetch_leave_dates_in_range(start=start, end=end)

    by_eid: Dict[str, EmployeeExportPivot] = {}
    for r in rows:
        eid = str(r.employee_id).strip()
        if not eid:
            continue
        wd = r.work_date or start
        if wd < start or wd > end:
            continue
        if eid not in by_eid:
            by_eid[eid] = EmployeeExportPivot(
                group_name=r.group_name,
                english_name=r.english_name,
                employee_id=eid,
            )
        on_leave = wd in leave_map.get(eid, set())
        by_eid[eid].daily_status[wd] = normalize_export_status(r, on_leave=on_leave)

    pivot = sorted(by_eid.values(), key=lambda x: x.employee_id)

    monthly_rest = absent = late = early = missed_punch = leave = normal_days = 0
    for p in pivot:
        for d in dates:
            st = p.daily_status.get(d, "")
            if not st:
                continue
            if st == "月休":
                monthly_rest += 1
            elif st == "请假":
                leave += 1
            elif st == "缺勤":
                absent += 1
            elif st == "缺卡":
                missed_punch += 1
            elif st == "迟到":
                late += 1
            elif st == "早退":
                early += 1
            elif st == "迟到+早退":
                late += 1
                early += 1
            elif st == "正常":
                normal_days += 1

    actual_count = sum(1 for p in pivot if any(p.daily_status.get(d) == "正常" for d in dates))

    overview = AttendanceExportOverview(
        expected_count=len(pivot),
        actual_count=actual_count,
        monthly_rest=monthly_rest,
        absent=absent,
        late=late,
        early=early,
        missed_punch=missed_punch,
        leave=leave,
    )
    return pivot, overview, dates


def _status_distribution_items(
    *,
    pivot: List[EmployeeExportPivot],
    dates: List[date],
    overview: AttendanceExportOverview,
) -> List[tuple[str, int, str]]:
    normal_days = sum(
        1 for p in pivot for d in dates if p.daily_status.get(d) == "正常"
    )
    values = {
        "正常": normal_days,
        "月休": overview.monthly_rest,
        "缺勤": overview.absent,
        "迟到": overview.late,
        "早退": overview.early,
        "缺卡": overview.missed_punch,
        "请假": overview.leave,
    }
    return [
        (name, values[name], color)
        for name, color in _STATUS_CHART_COLORS
        if values.get(name, 0) > 0
    ]


def _allocate_bar_segments(
    items: List[tuple[str, int, str]],
    *,
    width: int,
) -> List[tuple[str, int, str, int]]:
    total = sum(val for _name, val, _color in items)
    if total <= 0:
        return []
    segs: List[tuple[str, int, str, int]] = []
    used = 0
    for i, (name, val, color) in enumerate(items):
        if i == len(items) - 1:
            seg_w = width - used
        else:
            seg_w = max(1, round(val / total * width))
            remain_items = len(items) - i - 1
            seg_w = min(seg_w, width - used - remain_items)
            seg_w = max(1, seg_w)
        segs.append((name, val, color, seg_w))
        used += seg_w
    return segs


def _fill_merged_bar_segment(
    ws,
    *,
    row: int,
    start_col: int,
    end_col: int,
    color: str,
    label: str,
    center,
    label_font,
) -> None:
    from openpyxl.styles import PatternFill

    if end_col > start_col:
        ws.merge_cells(
            start_row=row,
            start_column=start_col,
            end_row=row,
            end_column=end_col,
        )
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    cell = ws.cell(row=row, column=start_col, value=label if end_col > start_col else "")
    cell.fill = fill
    cell.font = label_font
    cell.alignment = center


def _write_stacked_bar_row(
    ws,
    *,
    row: int,
    span_cols: int,
    items: List[tuple[str, int, str]],
    center,
    label_font,
) -> None:
    segs = _allocate_bar_segments(items, width=span_cols)
    col = 1
    for name, _val, color, seg_w in segs:
        if seg_w <= 0:
            continue
        seg_end = col + seg_w - 1
        _fill_merged_bar_segment(
            ws,
            row=row,
            start_col=col,
            end_col=seg_end,
            color=color,
            label=name if seg_w >= 2 else "",
            center=center,
            label_font=label_font,
        )
        col = seg_end + 1


def _write_mini_bar_row(
    ws,
    *,
    row: int,
    bar_start_col: int,
    bar_cols: int,
    color: str,
    ratio: float,
    center,
) -> None:
    from openpyxl.styles import PatternFill

    filled = max(1, round(ratio * bar_cols)) if ratio > 0 else 0
    empty_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    active_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    for i in range(bar_cols):
        col = bar_start_col + i
        cell = ws.cell(row=row, column=col, value="")
        cell.fill = active_fill if i < filled else empty_fill
        cell.alignment = center


def _write_color_block_distribution(
    ws,
    *,
    start_row: int,
    span_cols: int,
    items: List[tuple[str, int, str]],
    range_label: str,
    header_font,
    center,
) -> int:
    """纯表格色块条：全宽堆叠条 + 分项迷你条，无图表库。"""
    from openpyxl.styles import Font, PatternFill

    if not items:
        return start_row - 1

    total = sum(val for _name, val, _color in items)
    title_row = start_row
    stacked_bar_row = start_row + 1
    legend_header_row = start_row + 2
    mini_bar_start_col = 5
    mini_bar_cols = max(1, span_cols - mini_bar_start_col + 1)

    ws.merge_cells(
        start_row=title_row,
        start_column=1,
        end_row=title_row,
        end_column=span_cols,
    )
    title_cell = ws.cell(row=title_row, column=1, value=f"{range_label}状态分布")
    title_cell.font = header_font
    title_cell.alignment = center

    ws.row_dimensions[stacked_bar_row].height = 28
    bar_label_font = Font(bold=True, color="FFFFFF", size=9)
    _write_stacked_bar_row(
        ws,
        row=stacked_bar_row,
        span_cols=span_cols,
        items=items,
        center=center,
        label_font=bar_label_font,
    )

    legend_headers = ["色块", "状态", "人次", "占比", "图示"]
    for col_idx, header in enumerate(legend_headers, start=1):
        if col_idx > span_cols:
            break
        c = ws.cell(row=legend_header_row, column=col_idx, value=header)
        c.font = header_font
        c.alignment = center

    last_row = legend_header_row
    for name, val, color in items:
        row = last_row + 1
        last_row = row
        is_abnormal = name in {"缺勤", "缺卡", "迟到", "早退"}
        ratio = val / total if total else 0.0
        pct = f"{ratio * 100:.1f}%"

        swatch = ws.cell(row=row, column=1, value="")
        swatch.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        name_cell = ws.cell(row=row, column=2, value=name)
        name_cell.alignment = center
        if is_abnormal:
            name_cell.font = Font(bold=True, color="C00000")

        ws.cell(row=row, column=3, value=val).alignment = center
        ws.cell(row=row, column=4, value=pct).alignment = center
        _write_mini_bar_row(
            ws,
            row=row,
            bar_start_col=mini_bar_start_col,
            bar_cols=mini_bar_cols,
            color=color,
            ratio=ratio,
            center=center,
        )

    _apply_range_border(ws, min_row=title_row, max_row=last_row, min_col=1, max_col=span_cols)
    return last_row


def _apply_range_border(ws, *, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    from openpyxl.styles import Border, Side

    thin = Side(style="thin", color="000000")
    edge = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(row=r, column=c).border = edge


def build_attendance_export_grid(
    *,
    pivot: List[EmployeeExportPivot],
    dates: List[date],
    overview: AttendanceExportOverview,
    range_label: str,
    include_chart_section: bool = False,
) -> List[List[object]]:
    """与 XLSX 导出一致的数据网格（供 Google Sheets 等复用）。"""
    overview_headers = [
        "应出勤人数",
        "实出勤人数",
        "月休",
        "缺勤",
        "迟到",
        "早退",
        "缺卡",
        "请假",
    ]
    overview_values = [
        overview.expected_count,
        overview.actual_count,
        overview.monthly_rest,
        overview.absent,
        overview.late,
        overview.early,
        overview.missed_punch,
        overview.leave,
    ]
    grid: List[List[object]] = []
    title_row = [f"{range_label}数据概览"] + [""] * (len(overview_headers) - 1)
    grid.append(title_row)
    grid.append(overview_headers)
    grid.append(overview_values)

    if include_chart_section:
        chart_items = _status_distribution_items(pivot=pivot, dates=dates, overview=overview)
        if chart_items:
            grid.append([])
            grid.append([f"{range_label}状态分布", "状态", "人次", "占比"])
            total = sum(val for _name, val, _color in chart_items)
            for name, val, _color in chart_items:
                pct = f"{val / total * 100:.1f}%" if total else "0.0%"
                grid.append(["", name, val, pct])
            grid.append([])

    fixed_headers = ["群名", "员工", "工号"]
    date_headers = [_date_column_label(d) for d in dates]
    grid.append(fixed_headers + date_headers)
    fixed_cols = len(fixed_headers)
    for p in pivot:
        row: List[object] = [p.group_name, p.english_name, p.employee_id]
        for d in dates:
            row.append(p.daily_status.get(d, ""))
        grid.append(row)
    return grid


def abnormal_status_cells_in_export_grid(
    *,
    grid: List[List[object]],
    pivot: List[EmployeeExportPivot],
    dates: List[date],
) -> List[tuple[int, int]]:
    """返回需标黄的单元格（0-based row, col）。"""
    detail_header_idx: int | None = None
    for i, row in enumerate(grid):
        if row and str(row[0]).strip() == "群名":
            detail_header_idx = i
            break
    if detail_header_idx is None:
        return []
    fixed_cols = 3
    out: List[tuple[int, int]] = []
    for emp_i, p in enumerate(pivot):
        row_idx = detail_header_idx + 1 + emp_i
        for date_i, d in enumerate(dates):
            st = (p.daily_status.get(d) or "").strip()
            if st in _ABNORMAL_STATUSES:
                out.append((row_idx, fixed_cols + date_i))
    return out


def encode_attendance_export_xlsx(
    *,
    pivot: List[EmployeeExportPivot],
    dates: List[date],
    overview: AttendanceExportOverview,
    range_label: str,
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "考勤导出"

    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    overview_headers = [
        "应出勤人数",
        "实出勤人数",
        "月休",
        "缺勤",
        "迟到",
        "早退",
        "缺卡",
        "请假",
    ]
    overview_values = [
        overview.expected_count,
        overview.actual_count,
        overview.monthly_rest,
        overview.absent,
        overview.late,
        overview.early,
        overview.missed_punch,
        overview.leave,
    ]
    overview_cols = len(overview_headers)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=overview_cols)
    ws["A1"] = f"{range_label}数据概览"
    ws["A1"].font = header_font
    ws["A1"].alignment = center

    for col, h in enumerate(overview_headers, start=1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = header_font
        cell.alignment = center
        ws.cell(row=3, column=col, value=overview_values[col - 1]).alignment = center

    _apply_range_border(ws, min_row=1, max_row=3, min_col=1, max_col=overview_cols)

    chart_items = _status_distribution_items(pivot=pivot, dates=dates, overview=overview)
    distribution_end_row = 3
    if chart_items:
        distribution_end_row = _write_color_block_distribution(
            ws,
            start_row=5,
            span_cols=overview_cols,
            items=chart_items,
            range_label=range_label,
            header_font=header_font,
            center=center,
        )

    detail_start = distribution_end_row + 2
    fixed_headers = ["群名", "员工", "工号"]
    date_headers = [_date_column_label(d) for d in dates]
    detail_headers = fixed_headers + date_headers
    for col, h in enumerate(detail_headers, start=1):
        c = ws.cell(row=detail_start, column=col, value=h)
        c.font = header_font
        c.alignment = center

    fixed_cols = len(fixed_headers)
    for i, p in enumerate(pivot, start=detail_start + 1):
        ws.cell(row=i, column=1, value=p.group_name)
        ws.cell(row=i, column=2, value=p.english_name)
        ws.cell(row=i, column=3, value=p.employee_id)
        for j, d in enumerate(dates):
            col = fixed_cols + 1 + j
            st = p.daily_status.get(d, "")
            cell = ws.cell(row=i, column=col, value=st)
            cell.alignment = center
            if st in _ABNORMAL_STATUSES:
                cell.fill = yellow

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def today_in_tz(*, tz_name: str) -> date:
    from datetime import datetime

    return datetime.now(ZoneInfo(tz_name)).date()
