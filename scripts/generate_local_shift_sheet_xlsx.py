# -*- coding: utf-8 -*-
"""生成本地 6 月排班表 Excel（与统筹部 Google 模板同款布局）。"""
from __future__ import annotations

import argparse
from datetime import date
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]

# 统筹部模板顶部（图例 + 表头）
_TEMPLATE_TOP: list[list[str]] = [
    ["6月排班表"],
    [
        "4",
        "当月共有30天 员工需要出勤满26天才享有4天月休 ",
        "",
        "",
        "",
        "",
        "R",
        "00:00-09:00",
        "",
        "",
        "A",
        "07:00-16:00",
        "",
        "",
        "",
        "E",
        "11:00-20:00",
        "",
        "",
        "",
        "I",
        "15:00-00:00",
        "",
        "",
        "",
        "M",
        "19:00-04:00",
        "",
        "",
        "",
        "Q",
        "23:00-08:00",
        "",
        "",
        "出差",
        "",
        "",
        "⊗",
        "年假",
    ],
    [
        "",
        "",
        "",
        "",
        "",
        "",
        "S",
        "01:00-10:00",
        "",
        "",
        "B",
        "08:00-17:00",
        "",
        "",
        "",
        "F",
        "12:00-21:00",
        "",
        "",
        "",
        "J",
        "16:00-01:00",
        "",
        "",
        "",
        "N",
        "20:00-05:00",
        "",
        "",
        "",
        "",
        "新入职",
        "",
        "",
        "",
        "",
        "",
        "○",
        "请假",
    ],
    [
        "",
        "",
        "",
        "",
        "",
        "",
        "X",
        "05:00-14:00",
        "",
        "",
        "C",
        "09:00-18:00",
        "",
        "",
        "",
        "G",
        "13:00-22:00",
        "",
        "",
        "",
        "K",
        "17:00-02:00",
        "",
        "",
        "",
        "O",
        "21:00-06:00",
        "",
        "",
        "",
        "",
        "转岗",
        "",
        "",
        "",
        "",
        "",
        "▲",
        "休息",
    ],
    [
        "",
        "",
        "",
        "",
        "",
        "",
        "Y",
        "06:00-15:00",
        "",
        "",
        "D",
        "10:00-19:00",
        "",
        "",
        "",
        "H",
        "14:00-23:00",
        "",
        "",
        "",
        "L",
        "18:00-03:00",
        "",
        "",
        "",
        "P",
        "22:00-07:00",
        "",
        "",
        "",
        "",
        "离职",
        "",
        "",
        "",
        "",
        "",
        "※",
        "签证假",
    ],
    ["", "", "", "", "", "", "世界杯班次"],
    [
        "",
        "",
        "",
        "",
        "",
        "",
        "WR",
        "00:00-12:00",
        "",
        "",
        "WA",
        "07:00-19:00",
        "",
        "",
        "",
        "WE",
        "11:00-23:00",
        "",
        "",
        "",
        "WI",
        "15:00-03:00",
        "",
        "",
        "",
        "WM",
        "19:00-07:00",
        "",
        "",
        "",
        "WQ",
        "23:00-11:00",
        "",
        "",
        "出差",
        "",
        "",
        "⊗",
        "年假",
    ],
    [
        "",
        "",
        "",
        "",
        "",
        "",
        "WS",
        "01:00-13:00",
        "",
        "",
        "WB",
        "08:00-20:00",
        "",
        "",
        "",
        "WF",
        "12:00-00:00",
        "",
        "",
        "",
        "WJ",
        "16:00-04:00",
        "",
        "",
        "",
        "WN",
        "20:00-08:00",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "○",
        "请假",
    ],
    [
        "",
        "",
        "",
        "",
        "",
        "",
        "WX",
        "05:00-17:00",
        "",
        "",
        "WC",
        "09:00-21:00",
        "",
        "",
        "",
        "WG",
        "13:00-01:00",
        "",
        "",
        "",
        "WK",
        "17:00-05:00",
        "",
        "",
        "",
        "WO",
        "21:00-09:00",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "▲",
        "休息",
    ],
    [
        "",
        "",
        "",
        "",
        "",
        "",
        "WY",
        "06:00-18:00",
        "",
        "",
        "WD",
        "10:00-22:00",
        "",
        "",
        "",
        "WH",
        "14:00-02:00",
        "",
        "",
        "",
        "WL",
        "18:00-06:00",
        "",
        "",
        "",
        "WP",
        "22:00-10:00",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "※",
        "签证假",
    ],
    [
        "序号",
        "职位",
        "名字",
        "工号",
        "中文名",
        "入职时间",
        "一",
        "二",
        "三",
        "四",
        "五",
        "六",
        "日",
        "一",
        "二",
        "三",
        "四",
        "五",
        "六",
        "日",
        "一",
        "二",
        "三",
        "四",
        "五",
        "六",
        "日",
        "一",
        "二",
        "三",
        "四",
        "五",
        "六",
        "日",
        "一",
        "二",
        "日",
        "请假",
        "休息",
        "备注",
    ],
    [
        "",
        "",
        "",
        "",
        "",
        "",
        "1",
        "2",
        "3",
        "4",
        "5",
        "6",
        "7",
        "8",
        "9",
        "10",
        "11",
        "12",
        "13",
        "14",
        "15",
        "16",
        "17",
        "18",
        "19",
        "20",
        "21",
        "22",
        "23",
        "24",
        "25",
        "26",
        "27",
        "28",
        "29",
        "30",
        "1",
    ],
]

SHIFT_CODE = "WG"
REST_DAYS = frozenset({6, 13, 20, 27})

# 班次单元格底色（贴近 Google 模板）
_SHIFT_STYLE: dict[str, tuple[str, str]] = {
    "WG": ("843C0C", "FFFFFF"),
    "WO": ("2E75B6", "FFFFFF"),
    "WP": ("7030A0", "FFFFFF"),
    "WQ": ("4472C4", "FFFFFF"),
    "WH": ("C65911", "FFFFFF"),
    "WA": ("548235", "FFFFFF"),
    "WB": ("70AD47", "FFFFFF"),
    "▲": ("FFFFFF", "000000"),
}


def _weekend_days(year: int, month: int, *, days: int = 30) -> set[int]:
    out: set[int] = set()
    for d in range(1, days + 1):
        wd = date(year, month, d).weekday()
        if wd >= 5:
            out.add(d)
    return out


def _build_employee_row(
    *,
    seq: int,
    english_name: str,
    employee_id: str,
    chinese_name: str = "",
    join_date: str = "2026-6-1",
) -> list[str]:
    row = [""] * 40
    row[0] = str(seq)
    row[1] = "UX设计师"
    row[2] = english_name
    row[3] = employee_id
    row[4] = chinese_name
    row[5] = join_date
    for day in range(1, 31):
        col = 5 + day
        row[col] = "▲" if day in REST_DAYS else SHIFT_CODE
    row[37] = "0"
    row[38] = str(len(REST_DAYS))
    return row


def generate_local_shift_xlsx(
    *,
    out_path: Path,
    employees: list[tuple[str, str, str]],
    year_month: str = "2026-06",
) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    y, m = map(int, year_month.split("-"))
    weekend_days = _weekend_days(y, m)

    emp_rows = [
        _build_employee_row(seq=i + 1, english_name=name, employee_id=eid, chinese_name=cn)
        for i, (name, eid, cn) in enumerate(employees)
    ]
    footer = [
        ["", "", "", "", str(len(employees)), "UX设计组", str(len(employees)), str(len(employees))],
        ["", "", "", "", "0", "夏荷组", "0", "0"],
        ["", "", "", "", str(len(employees)), "统筹部总计", str(len(employees)), str(len(employees))],
    ]
    all_rows: list[list[str]] = (
        _TEMPLATE_TOP + [["UX设计组"]] + emp_rows + [[]] + footer
    )

    wb = Workbook()
    ws = wb.active
    ws.title = f"排班 {year_month}"

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    title_font = Font(bold=True, size=16)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    weekend_fill = PatternFill("solid", fgColor="FFF2CC")
    group_fill = PatternFill("solid", fgColor="E2EFDA")

    max_col = max(len(r) for r in all_rows)
    for r_idx, row in enumerate(all_rows, start=1):
        for c_idx in range(1, max_col + 1):
            val = row[c_idx - 1] if c_idx - 1 < len(row) else ""
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = center
            cell.border = border

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws["A1"].font = title_font

    header_row = 11
    date_row = 12
    for c in range(7, 37):
        ws.cell(row=header_row, column=c).fill = header_fill
        ws.cell(row=header_row, column=c).font = Font(bold=True)
        ws.cell(row=date_row, column=c).fill = header_fill
        ws.cell(row=date_row, column=c).font = Font(bold=True)

    for day in range(1, 31):
        col = 6 + day
        if day in weekend_days:
            ws.cell(row=header_row, column=col).fill = weekend_fill
            ws.cell(row=date_row, column=col).fill = weekend_fill

    group_row = 13
    ws.merge_cells(start_row=group_row, start_column=1, end_row=group_row, end_column=6)
    ws.cell(row=group_row, column=1).fill = group_fill
    ws.cell(row=group_row, column=1).font = Font(bold=True)

    for r_idx in range(group_row + 1, group_row + 1 + len(emp_rows)):
        for day in range(1, 31):
            col = 6 + day
            code = str(ws.cell(row=r_idx, column=col).value or "").strip().upper()
            bg, fg = _SHIFT_STYLE.get(code, ("FFFFFF", "000000"))
            cell = ws.cell(row=r_idx, column=col)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.font = Font(color=fg, bold=True)
            if day in weekend_days:
                pass

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[header_row].height = 22
    ws.row_dimensions[date_row].height = 20
    for c in range(1, 7):
        ws.column_dimensions[get_column_letter(c)].width = 12
    for c in range(7, 37):
        ws.column_dimensions[get_column_letter(c)].width = 4.5

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def main() -> int:
    parser = argparse.ArgumentParser(description="生成本地排班表 Excel（统筹部模板同款）")
    parser.add_argument(
        "--out",
        type=Path,
        default=ROOT.parent / "排班2026-06_99999_bodyhh.xlsx",
        help="输出 xlsx 路径",
    )
    args = parser.parse_args()

    employees = [
        ("nliliii", "99999", ""),
        ("bodyhh", "102", ""),
    ]
    path = generate_local_shift_xlsx(out_path=args.out.resolve(), employees=employees)
    print(f"已生成: {path}")
    print(f"员工: 99999/nliliii, 102/bodyhh | 班次: WG (13:00-01:00)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
